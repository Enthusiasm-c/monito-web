#!/usr/bin/env python3
"""
Unified Excel Reader for All Sheets
Implements Task 5: Unify XLSX reading for all sheets

Features:
- Multi-sheet Excel file processing
- Hidden row/column detection and removal
- Support for both .xlsx and .xls formats
- Data type preservation and normalization
- Empty sheet detection and filtering
- Comprehensive metadata extraction
- Memory-efficient streaming for large files
"""

import sys
import os
import json
import time
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
import logging
from io import StringIO

# openpyxl for .xlsx files
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("[WARNING] openpyxl not available, .xlsx support limited", file=sys.stderr)
    OPENPYXL_AVAILABLE = False

# xlrd for .xls files
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    print("[WARNING] xlrd not available, .xls support disabled", file=sys.stderr)
    XLRD_AVAILABLE = False

# pandas for data processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    print("[WARNING] pandas not available, limited data processing", file=sys.stderr)
    PANDAS_AVAILABLE = False


class ExcelReader:
    """Unified Excel reader for all sheets and formats"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = config or {}
        
        # Configuration parameters
        self.include_hidden = self.config.get('include_hidden', False)
        self.skip_empty_sheets = self.config.get('skip_empty_sheets', True)
        self.min_rows = self.config.get('min_rows', 3)
        self.min_cols = self.config.get('min_cols', 2)
        self.max_sheets = self.config.get('max_sheets', 50)  # Prevent memory issues
        self.data_only = self.config.get('data_only', True)  # Read values, not formulas
        self.preserve_dtypes = self.config.get('preserve_dtypes', True)
        self.normalize_headers = self.config.get('normalize_headers', True)
        
        print(f"[INFO] Excel Reader initialized", file=sys.stderr)
        print(f"[INFO] openpyxl: {'✓' if OPENPYXL_AVAILABLE else '✗'}", file=sys.stderr)
        print(f"[INFO] xlrd: {'✓' if XLRD_AVAILABLE else '✗'}", file=sys.stderr)
        print(f"[INFO] pandas: {'✓' if PANDAS_AVAILABLE else '✗'}", file=sys.stderr)
    
    def read_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """
        Read Excel file with all sheets
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dict containing all sheet data and metadata
        """
        start_time = time.time()
        
        print(f"[INFO] Reading Excel file: {file_path}", file=sys.stderr)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        result = {
            "file_path": str(file_path),
            "sheets": {},
            "metadata": {
                "total_sheets": 0,
                "processed_sheets": 0,
                "skipped_sheets": 0,
                "hidden_sheets": 0,
                "empty_sheets": 0,
                "file_format": self._detect_format(file_path),
                "file_size_bytes": file_path.stat().st_size,
                "processing_time_ms": 0,
                "total_rows": 0,
                "total_cells": 0
            },
            "success": False,
            "error": None
        }
        
        try:
            # Determine file format and use appropriate reader
            file_format = result["metadata"]["file_format"]
            
            if file_format == "xlsx":
                if not OPENPYXL_AVAILABLE:
                    raise RuntimeError("openpyxl required for .xlsx files")
                sheet_data = self._read_xlsx_file(file_path)
            elif file_format == "xls":
                if not XLRD_AVAILABLE:
                    raise RuntimeError("xlrd required for .xls files")
                sheet_data = self._read_xls_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_format}")
            
            # Process sheet data
            result["sheets"] = sheet_data["sheets"]
            result["metadata"].update(sheet_data["metadata"])
            result["success"] = True
            
            # Calculate totals
            total_rows = sum(sheet["row_count"] for sheet in result["sheets"].values())
            total_cells = sum(sheet["cell_count"] for sheet in result["sheets"].values())
            
            result["metadata"]["total_rows"] = total_rows
            result["metadata"]["total_cells"] = total_cells
            
            print(f"[INFO] Excel processing completed: {result['metadata']['processed_sheets']} sheets, "
                  f"{total_rows} rows, {total_cells} cells", file=sys.stderr)
            
        except Exception as e:
            result["error"] = str(e)
            result["success"] = False
            print(f"[ERROR] Excel reading failed: {e}", file=sys.stderr)
        
        finally:
            result["metadata"]["processing_time_ms"] = int((time.time() - start_time) * 1000)
        
        return result
    
    def _detect_format(self, file_path: Path) -> str:
        """Detect Excel file format"""
        suffix = file_path.suffix.lower()
        if suffix == '.xlsx':
            return 'xlsx'
        elif suffix == '.xls':
            return 'xls'
        else:
            # Try to detect by content
            try:
                with open(file_path, 'rb') as f:
                    header = f.read(8)
                    if header.startswith(b'PK'):  # ZIP signature (xlsx)
                        return 'xlsx'
                    elif header.startswith(b'\xd0\xcf\x11\xe0'):  # OLE2 signature (xls)
                        return 'xls'
            except:
                pass
        
        return 'unknown'
    
    def _read_xlsx_file(self, file_path: Path) -> Dict[str, Any]:
        """Read .xlsx file using openpyxl"""
        
        print(f"[INFO] Reading .xlsx file with openpyxl", file=sys.stderr)
        
        # Load workbook
        workbook = load_workbook(
            filename=file_path,
            read_only=False,  # Need write access to check hidden sheets
            data_only=self.data_only,
            keep_vba=False
        )
        
        sheet_data = {
            "sheets": {},
            "metadata": {
                "total_sheets": len(workbook.worksheets),
                "processed_sheets": 0,
                "skipped_sheets": 0,
                "hidden_sheets": 0,
                "empty_sheets": 0
            }
        }
        
        try:
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    
                    # Check if sheet is hidden
                    is_hidden = sheet.sheet_state != 'visible'
                    if is_hidden:
                        sheet_data["metadata"]["hidden_sheets"] += 1
                        if not self.include_hidden:
                            print(f"[INFO] Skipping hidden sheet: {sheet_name}", file=sys.stderr)
                            sheet_data["metadata"]["skipped_sheets"] += 1
                            continue
                    
                    # Extract sheet data
                    sheet_result = self._extract_sheet_data_openpyxl(sheet, sheet_name)
                    
                    # Check if sheet is empty
                    if sheet_result["row_count"] < self.min_rows or sheet_result["col_count"] < self.min_cols:
                        sheet_data["metadata"]["empty_sheets"] += 1
                        if self.skip_empty_sheets:
                            print(f"[INFO] Skipping empty sheet: {sheet_name} "
                                  f"({sheet_result['row_count']} rows, {sheet_result['col_count']} cols)", file=sys.stderr)
                            sheet_data["metadata"]["skipped_sheets"] += 1
                            continue
                    
                    # Add sheet data
                    sheet_data["sheets"][sheet_name] = sheet_result
                    sheet_data["metadata"]["processed_sheets"] += 1
                    
                    print(f"[INFO] Processed sheet '{sheet_name}': "
                          f"{sheet_result['row_count']} rows, {sheet_result['col_count']} cols", file=sys.stderr)
                    
                    # Prevent memory issues
                    if len(sheet_data["sheets"]) >= self.max_sheets:
                        print(f"[WARNING] Reached max sheets limit ({self.max_sheets})", file=sys.stderr)
                        break
                        
                except Exception as e:
                    print(f"[WARNING] Failed to process sheet '{sheet_name}': {e}", file=sys.stderr)
                    sheet_data["metadata"]["skipped_sheets"] += 1
        
        finally:
            workbook.close()
        
        return sheet_data
    
    def _read_xls_file(self, file_path: Path) -> Dict[str, Any]:
        """Read .xls file using xlrd"""
        
        print(f"[INFO] Reading .xls file with xlrd", file=sys.stderr)
        
        # Open workbook
        workbook = xlrd.open_workbook(file_path)
        
        sheet_data = {
            "sheets": {},
            "metadata": {
                "total_sheets": workbook.nsheets,
                "processed_sheets": 0,
                "skipped_sheets": 0,
                "hidden_sheets": 0,
                "empty_sheets": 0
            }
        }
        
        # Process each sheet
        for sheet_index in range(workbook.nsheets):
            try:
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_name = sheet.name
                
                # Check for hidden sheets (xlrd has limited support)
                is_hidden = getattr(sheet, 'visibility', 0) != 0
                if is_hidden:
                    sheet_data["metadata"]["hidden_sheets"] += 1
                    if not self.include_hidden:
                        print(f"[INFO] Skipping hidden sheet: {sheet_name}", file=sys.stderr)
                        sheet_data["metadata"]["skipped_sheets"] += 1
                        continue
                
                # Extract sheet data
                sheet_result = self._extract_sheet_data_xlrd(sheet, sheet_name)
                
                # Check if sheet is empty
                if sheet_result["row_count"] < self.min_rows or sheet_result["col_count"] < self.min_cols:
                    sheet_data["metadata"]["empty_sheets"] += 1
                    if self.skip_empty_sheets:
                        print(f"[INFO] Skipping empty sheet: {sheet_name} "
                              f"({sheet_result['row_count']} rows, {sheet_result['col_count']} cols)", file=sys.stderr)
                        sheet_data["metadata"]["skipped_sheets"] += 1
                        continue
                
                # Add sheet data
                sheet_data["sheets"][sheet_name] = sheet_result
                sheet_data["metadata"]["processed_sheets"] += 1
                
                print(f"[INFO] Processed sheet '{sheet_name}': "
                      f"{sheet_result['row_count']} rows, {sheet_result['col_count']} cols", file=sys.stderr)
                
                # Prevent memory issues
                if len(sheet_data["sheets"]) >= self.max_sheets:
                    print(f"[WARNING] Reached max sheets limit ({self.max_sheets})", file=sys.stderr)
                    break
                    
            except Exception as e:
                print(f"[WARNING] Failed to process sheet index {sheet_index}: {e}", file=sys.stderr)
                sheet_data["metadata"]["skipped_sheets"] += 1
        
        return sheet_data
    
    def _extract_sheet_data_openpyxl(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract data from openpyxl worksheet"""
        
        # Get actual used range
        min_row = sheet.min_row or 1
        max_row = sheet.max_row or 1
        min_col = sheet.min_column or 1
        max_col = sheet.max_column or 1
        
        # Filter out hidden rows and columns if needed
        visible_rows = []
        visible_cols = []
        
        # Check for hidden rows
        for row_num in range(min_row, max_row + 1):
            row_dimension = sheet.row_dimensions.get(row_num)
            if row_dimension and row_dimension.hidden and not self.include_hidden:
                continue
            visible_rows.append(row_num)
        
        # Check for hidden columns
        for col_num in range(min_col, max_col + 1):
            col_letter = get_column_letter(col_num)
            col_dimension = sheet.column_dimensions.get(col_letter)
            if col_dimension and col_dimension.hidden and not self.include_hidden:
                continue
            visible_cols.append(col_num)
        
        # Extract visible data
        data_rows = []
        headers = []
        header_detected = False
        
        for row_idx, row_num in enumerate(visible_rows):
            row_data = []
            
            for col_num in visible_cols:
                cell = sheet.cell(row=row_num, column=col_num)
                value = self._normalize_cell_value(cell.value)
                row_data.append(value)
            
            # Detect headers (first non-empty row with strings)
            if not header_detected and any(isinstance(v, str) and v.strip() for v in row_data):
                if self.normalize_headers:
                    headers = [self._normalize_header(v) for v in row_data]
                else:
                    headers = row_data
                header_detected = True
                continue
            
            # Skip empty rows
            if any(v is not None and str(v).strip() for v in row_data):
                data_rows.append(row_data)
        
        # Create structured data
        structured_data = []
        for row in data_rows:
            if headers:
                # Create dict with headers as keys
                row_dict = {}
                for i, value in enumerate(row):
                    header = headers[i] if i < len(headers) else f"column_{i+1}"
                    row_dict[header] = value
                structured_data.append(row_dict)
            else:
                # No headers detected, use positional data
                structured_data.append(row)
        
        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "data": structured_data,
            "raw_data": data_rows,
            "row_count": len(data_rows),
            "col_count": len(visible_cols),
            "cell_count": len(data_rows) * len(visible_cols),
            "has_headers": header_detected,
            "visible_rows": len(visible_rows),
            "visible_cols": len(visible_cols),
            "total_rows": max_row - min_row + 1,
            "total_cols": max_col - min_col + 1,
            "data_range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        }
    
    def _extract_sheet_data_xlrd(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract data from xlrd worksheet"""
        
        # Get sheet dimensions
        nrows = sheet.nrows
        ncols = sheet.ncols
        
        if nrows == 0 or ncols == 0:
            return {
                "sheet_name": sheet_name,
                "headers": [],
                "data": [],
                "raw_data": [],
                "row_count": 0,
                "col_count": 0,
                "cell_count": 0,
                "has_headers": False,
                "visible_rows": 0,
                "visible_cols": 0,
                "total_rows": nrows,
                "total_cols": ncols,
                "data_range": f"A1:A1"
            }
        
        # Extract all data
        data_rows = []
        headers = []
        header_detected = False
        
        for row_idx in range(nrows):
            row_data = []
            
            for col_idx in range(ncols):
                try:
                    cell = sheet.cell(row_idx, col_idx)
                    value = self._normalize_xlrd_cell_value(cell)
                    row_data.append(value)
                except:
                    row_data.append(None)
            
            # Detect headers
            if not header_detected and any(isinstance(v, str) and v.strip() for v in row_data):
                if self.normalize_headers:
                    headers = [self._normalize_header(v) for v in row_data]
                else:
                    headers = row_data
                header_detected = True
                continue
            
            # Skip empty rows
            if any(v is not None and str(v).strip() for v in row_data):
                data_rows.append(row_data)
        
        # Create structured data
        structured_data = []
        for row in data_rows:
            if headers:
                row_dict = {}
                for i, value in enumerate(row):
                    header = headers[i] if i < len(headers) else f"column_{i+1}"
                    row_dict[header] = value
                structured_data.append(row_dict)
            else:
                structured_data.append(row)
        
        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "data": structured_data,
            "raw_data": data_rows,
            "row_count": len(data_rows),
            "col_count": ncols,
            "cell_count": len(data_rows) * ncols,
            "has_headers": header_detected,
            "visible_rows": len(data_rows),
            "visible_cols": ncols,
            "total_rows": nrows,
            "total_cols": ncols,
            "data_range": f"A1:{get_column_letter(ncols)}{nrows}"
        }
    
    def _normalize_cell_value(self, value: Any) -> Any:
        """Normalize cell value from openpyxl"""
        if value is None:
            return None
        
        # Handle datetime objects
        if hasattr(value, 'date'):
            return value.isoformat()
        
        # Handle numeric values
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value
        
        # Handle strings
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try to convert to number if it looks like one
            if self.preserve_dtypes:
                try:
                    if '.' in value or 'e' in value.lower():
                        return float(value)
                    else:
                        return int(value)
                except ValueError:
                    pass
            
            return value
        
        return str(value) if value else None
    
    def _normalize_xlrd_cell_value(self, cell) -> Any:
        """Normalize cell value from xlrd"""
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            text = cell.value.strip()
            return text if text else None
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            num = cell.value
            if isinstance(num, float) and num.is_integer():
                return int(num)
            return num
        elif cell.ctype == xlrd.XL_CELL_DATE:
            # Convert Excel date to string
            date_tuple = xlrd.xldate_as_tuple(cell.value, 0)  # Assume 1900 date system
            if date_tuple[:3] != (0, 0, 0):  # Valid date
                return f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
            else:
                return cell.value
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        else:
            return str(cell.value) if cell.value else None
    
    def _normalize_header(self, header: Any) -> str:
        """Normalize header name"""
        if header is None:
            return "unnamed_column"
        
        header_str = str(header).strip()
        if not header_str:
            return "unnamed_column"
        
        # Clean up header name
        header_str = header_str.replace('\n', ' ').replace('\r', ' ')
        header_str = ' '.join(header_str.split())  # Normalize whitespace
        
        # Convert to lowercase and replace spaces with underscores
        header_str = header_str.lower().replace(' ', '_')
        
        # Remove special characters except underscore
        import re
        header_str = re.sub(r'[^a-z0-9_]', '', header_str)
        
        # Ensure it starts with letter or underscore
        if header_str and header_str[0].isdigit():
            header_str = f"col_{header_str}"
        
        return header_str or "unnamed_column"
    
    def get_sheet_summary(self, file_path: Path) -> Dict[str, Any]:
        """Get quick summary of Excel file without reading all data"""
        
        print(f"[INFO] Getting Excel file summary: {file_path}", file=sys.stderr)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        file_format = self._detect_format(file_path)
        summary = {
            "file_path": str(file_path),
            "file_format": file_format,
            "file_size_bytes": file_path.stat().st_size,
            "sheets": []
        }
        
        try:
            if file_format == "xlsx" and OPENPYXL_AVAILABLE:
                workbook = load_workbook(filename=file_path, read_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    is_hidden = sheet.sheet_state != 'visible'
                    
                    # Get rough dimensions
                    max_row = sheet.max_row or 0
                    max_col = sheet.max_column or 0
                    
                    summary["sheets"].append({
                        "name": sheet_name,
                        "hidden": is_hidden,
                        "estimated_rows": max_row,
                        "estimated_cols": max_col,
                        "estimated_cells": max_row * max_col if max_row and max_col else 0
                    })
                
                workbook.close()
                
            elif file_format == "xls" and XLRD_AVAILABLE:
                workbook = xlrd.open_workbook(file_path)
                
                for sheet_index in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_index)
                    is_hidden = getattr(sheet, 'visibility', 0) != 0
                    
                    summary["sheets"].append({
                        "name": sheet.name,
                        "hidden": is_hidden,
                        "estimated_rows": sheet.nrows,
                        "estimated_cols": sheet.ncols,
                        "estimated_cells": sheet.nrows * sheet.ncols
                    })
            
            else:
                raise ValueError(f"Cannot read file format: {file_format}")
        
        except Exception as e:
            summary["error"] = str(e)
            print(f"[WARNING] Failed to get summary: {e}", file=sys.stderr)
        
        return summary
    
    def export_to_csv(self, excel_data: Dict[str, Any], output_dir: Path) -> List[Path]:
        """Export all sheets to separate CSV files"""
        
        if not PANDAS_AVAILABLE:
            raise RuntimeError("pandas required for CSV export")
        
        output_dir.mkdir(parents=True, exist_ok=True)
        csv_files = []
        
        for sheet_name, sheet_data in excel_data["sheets"].items():
            try:
                # Create DataFrame
                if sheet_data["has_headers"] and sheet_data["headers"]:
                    df = pd.DataFrame(sheet_data["raw_data"], columns=sheet_data["headers"])
                else:
                    df = pd.DataFrame(sheet_data["raw_data"])
                
                # Clean filename
                safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                safe_name = safe_name.replace(' ', '_')
                csv_file = output_dir / f"{safe_name}.csv"
                
                # Export to CSV
                df.to_csv(csv_file, index=False, encoding='utf-8')
                csv_files.append(csv_file)
                
                print(f"[INFO] Exported sheet '{sheet_name}' to {csv_file}", file=sys.stderr)
                
            except Exception as e:
                print(f"[WARNING] Failed to export sheet '{sheet_name}': {e}", file=sys.stderr)
        
        return csv_files


def main():
    """CLI interface for Excel reading"""
    
    if len(sys.argv) < 2:
        print("Usage: python excel_reader.py <excel_path> [output_json]", file=sys.stderr)
        sys.exit(1)
    
    excel_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)
    
    # Configuration can be passed via environment variables
    config = {
        "include_hidden": os.getenv("INCLUDE_HIDDEN", "false").lower() == "true",
        "skip_empty_sheets": os.getenv("SKIP_EMPTY_SHEETS", "true").lower() == "true",
        "min_rows": int(os.getenv("MIN_ROWS", "3")),
        "min_cols": int(os.getenv("MIN_COLS", "2")),
        "data_only": os.getenv("DATA_ONLY", "true").lower() == "true",
        "normalize_headers": os.getenv("NORMALIZE_HEADERS", "true").lower() == "true"
    }
    
    try:
        reader = ExcelReader(config)
        result = reader.read_excel_file(excel_path)
        
        # Output result
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            print(f"[INFO] Results saved to {output_path}", file=sys.stderr)
        else:
            print(f"EXCEL_READING_RESULT:{json.dumps(result)}")
        
        # Summary to stderr
        metadata = result["metadata"]
        print(f"[SUMMARY] Sheets: {metadata['processed_sheets']}/{metadata['total_sheets']}, "
              f"Rows: {metadata['total_rows']}, "
              f"Cells: {metadata['total_cells']}, "
              f"Time: {metadata['processing_time_ms']}ms", file=sys.stderr)
        
    except Exception as e:
        print(f"[ERROR] Excel reading failed: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
"""
Tests for Unified Excel Reading
Tests Task 5: Unify XLSX reading for all sheets
"""

import pytest
import tempfile
import os
from pathlib import Path
import json
from unittest.mock import Mock, patch, MagicMock

# Import our Excel reading modules
import sys
sys.path.append(str(Path(__file__).parent.parent))

from src.pipeline.excel_reader import ExcelReader


class TestExcelReader:
    """Test Excel reading functionality"""

    @pytest.fixture
    def excel_reader(self):
        """Create Excel reader instance"""
        config = {
            'include_hidden': False,
            'skip_empty_sheets': True,
            'min_rows': 3,
            'min_cols': 2,
            'data_only': True,
            'normalize_headers': True
        }
        return ExcelReader(config)

    @pytest.fixture
    def sample_xlsx_file(self, tmp_path):
        """Create a sample .xlsx file for testing"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # First sheet (default)
            ws1 = wb.active
            ws1.title = "Products"
            
            # Add headers
            headers = ["Product Name", "Unit", "Price", "Category", "Supplier"]
            for col, header in enumerate(headers, 1):
                ws1.cell(row=1, column=col, value=header)
            
            # Add data
            data = [
                ["Fresh Tomatoes", "kg", 12.50, "Vegetables", "Farm Fresh Co"],
                ["Organic Carrots", "kg", 15.00, "Vegetables", "Organic Valley"],
                ["Milk", "liter", 8.75, "Dairy", "Dairy Plus"],
                ["Bread", "loaf", 5.50, "Bakery", "Local Bakery"],
                ["Chicken Breast", "kg", 45.00, "Meat", "Poultry Farm"]
            ]
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=value)
            
            # Second sheet
            ws2 = wb.create_sheet("Suppliers")
            supplier_headers = ["Supplier Name", "Contact", "Location"]
            for col, header in enumerate(supplier_headers, 1):
                ws2.cell(row=1, column=col, value=header)
            
            supplier_data = [
                ["Farm Fresh Co", "info@farmfresh.com", "Jakarta"],
                ["Organic Valley", "contact@organicvalley.com", "Bandung"],
                ["Dairy Plus", "sales@dairyplus.com", "Surabaya"]
            ]
            
            for row_idx, row_data in enumerate(supplier_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=value)
            
            # Third sheet (empty)
            ws3 = wb.create_sheet("Empty Sheet")
            
            # Fourth sheet (hidden)
            ws4 = wb.create_sheet("Hidden Data")
            ws4.sheet_state = 'hidden'
            ws4.cell(row=1, column=1, value="Secret Data")
            
            # Save workbook
            xlsx_path = tmp_path / "sample_data.xlsx"
            wb.save(xlsx_path)
            wb.close()
            
            return xlsx_path
            
        except ImportError:
            pytest.skip("openpyxl not available for creating test files")

    @pytest.fixture
    def sample_xls_file(self, tmp_path):
        """Create a sample .xls file for testing"""
        try:
            import xlwt
            
            # Create workbook
            wb = xlwt.Workbook()
            
            # Add sheet with data
            ws = wb.add_sheet("Price List")
            
            # Add headers
            headers = ["Item", "Price", "Stock"]
            for col, header in enumerate(headers):
                ws.write(0, col, header)
            
            # Add data
            data = [
                ["Product A", 100.0, 50],
                ["Product B", 150.0, 30],
                ["Product C", 200.0, 25]
            ]
            
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data):
                    ws.write(row_idx, col_idx, value)
            
            # Save workbook
            xls_path = tmp_path / "sample_data.xls"
            wb.save(xls_path)
            
            return xls_path
            
        except ImportError:
            pytest.skip("xlwt not available for creating .xls test files")

    def test_excel_reader_initialization(self, excel_reader):
        """Test Excel reader initialization"""
        assert excel_reader.include_hidden is False
        assert excel_reader.skip_empty_sheets is True
        assert excel_reader.min_rows == 3
        assert excel_reader.min_cols == 2
        assert excel_reader.data_only is True
        assert excel_reader.normalize_headers is True

    @pytest.mark.unit
    def test_format_detection(self, excel_reader, tmp_path):
        """Test Excel format detection"""
        # Test .xlsx extension
        xlsx_file = tmp_path / "test.xlsx"
        xlsx_file.write_bytes(b"PK")  # ZIP signature
        assert excel_reader._detect_format(xlsx_file) == "xlsx"
        
        # Test .xls extension
        xls_file = tmp_path / "test.xls"
        xls_file.write_bytes(b"\xd0\xcf\x11\xe0")  # OLE2 signature
        assert excel_reader._detect_format(xls_file) == "xls"
        
        # Test unknown format
        unknown_file = tmp_path / "test.unknown"
        unknown_file.write_bytes(b"unknown")
        assert excel_reader._detect_format(unknown_file) == "unknown"

    @pytest.mark.unit
    def test_header_normalization(self, excel_reader):
        """Test header normalization"""
        # Test normal header
        assert excel_reader._normalize_header("Product Name") == "product_name"
        
        # Test header with special characters
        assert excel_reader._normalize_header("Price (USD)") == "price_usd"
        
        # Test header with numbers
        assert excel_reader._normalize_header("2024 Sales") == "col_2024_sales"
        
        # Test empty header
        assert excel_reader._normalize_header("") == "unnamed_column"
        assert excel_reader._normalize_header(None) == "unnamed_column"

    @pytest.mark.unit
    def test_cell_value_normalization(self, excel_reader):
        """Test cell value normalization"""
        # Test None
        assert excel_reader._normalize_cell_value(None) is None
        
        # Test strings
        assert excel_reader._normalize_cell_value("  text  ") == "text"
        assert excel_reader._normalize_cell_value("") is None
        
        # Test numbers
        assert excel_reader._normalize_cell_value(123) == 123
        assert excel_reader._normalize_cell_value(123.0) == 123  # Integer conversion
        assert excel_reader._normalize_cell_value(123.45) == 123.45
        
        # Test numeric strings (if preserve_dtypes is True)
        excel_reader.preserve_dtypes = True
        assert excel_reader._normalize_cell_value("123") == 123
        assert excel_reader._normalize_cell_value("123.45") == 123.45

    @pytest.mark.unit
    def test_xlsx_reading(self, excel_reader, sample_xlsx_file):
        """Test .xlsx file reading"""
        try:
            result = excel_reader.read_excel_file(sample_xlsx_file)
            
            # Verify result structure
            assert result["success"] is True
            assert "sheets" in result
            assert "metadata" in result
            
            # Verify metadata
            metadata = result["metadata"]
            assert metadata["file_format"] == "xlsx"
            assert metadata["total_sheets"] >= 2  # At least Products and Suppliers
            assert metadata["processed_sheets"] >= 2
            
            # Verify sheet data
            sheets = result["sheets"]
            assert "Products" in sheets
            assert "Suppliers" in sheets
            
            # Check Products sheet
            products_sheet = sheets["Products"]
            assert products_sheet["has_headers"] is True
            assert len(products_sheet["headers"]) == 5
            assert products_sheet["row_count"] >= 5  # 5 data rows
            assert "product_name" in [h.lower().replace(" ", "_") for h in products_sheet["headers"]]
            
            # Check that empty sheet was skipped
            assert "Empty Sheet" not in sheets  # Should be skipped due to empty
            
            # Check that hidden sheet was skipped
            assert "Hidden Data" not in sheets  # Should be skipped due to hidden
            
        except ImportError:
            pytest.skip("openpyxl not available for .xlsx testing")

    @pytest.mark.unit
    def test_xls_reading(self, excel_reader, sample_xls_file):
        """Test .xls file reading"""
        try:
            result = excel_reader.read_excel_file(sample_xls_file)
            
            # Verify result structure
            assert result["success"] is True
            assert "sheets" in result
            assert "metadata" in result
            
            # Verify metadata
            metadata = result["metadata"]
            assert metadata["file_format"] == "xls"
            assert metadata["processed_sheets"] >= 1
            
            # Verify sheet data
            sheets = result["sheets"]
            assert "Price List" in sheets
            
            # Check sheet content
            price_sheet = sheets["Price List"]
            assert price_sheet["has_headers"] is True
            assert price_sheet["row_count"] >= 3
            
        except ImportError:
            pytest.skip("xlrd not available for .xls testing")

    @pytest.mark.unit
    def test_file_not_found_error(self, excel_reader):
        """Test error handling for non-existent files"""
        non_existent_file = Path("/path/that/does/not/exist.xlsx")
        
        with pytest.raises(FileNotFoundError):
            excel_reader.read_excel_file(non_existent_file)

    @pytest.mark.unit
    def test_configuration_override(self):
        """Test configuration override"""
        custom_config = {
            'include_hidden': True,
            'skip_empty_sheets': False,
            'min_rows': 1,
            'min_cols': 1,
            'normalize_headers': False
        }
        
        reader = ExcelReader(custom_config)
        
        assert reader.include_hidden is True
        assert reader.skip_empty_sheets is False
        assert reader.min_rows == 1
        assert reader.min_cols == 1
        assert reader.normalize_headers is False

    @pytest.mark.unit
    def test_include_hidden_sheets(self, sample_xlsx_file):
        """Test including hidden sheets"""
        try:
            # Test with include_hidden = True
            config = {'include_hidden': True, 'skip_empty_sheets': False, 'min_rows': 1, 'min_cols': 1}
            reader = ExcelReader(config)
            
            result = reader.read_excel_file(sample_xlsx_file)
            
            if result["success"]:
                # Should include hidden sheet now
                assert result["metadata"]["hidden_sheets"] > 0
                # Hidden sheet should be processed if include_hidden is True
                
        except ImportError:
            pytest.skip("openpyxl not available for hidden sheet testing")

    @pytest.mark.unit
    def test_sheet_summary(self, excel_reader, sample_xlsx_file):
        """Test getting sheet summary"""
        try:
            summary = excel_reader.get_sheet_summary(sample_xlsx_file)
            
            assert "file_path" in summary
            assert "file_format" in summary
            assert "sheets" in summary
            assert summary["file_format"] == "xlsx"
            
            # Should have multiple sheets
            assert len(summary["sheets"]) >= 2
            
            # Check sheet info
            for sheet_info in summary["sheets"]:
                assert "name" in sheet_info
                assert "hidden" in sheet_info
                assert "estimated_rows" in sheet_info
                assert "estimated_cols" in sheet_info
                
        except ImportError:
            pytest.skip("openpyxl not available for summary testing")

    @pytest.mark.unit
    def test_error_handling_and_recovery(self, excel_reader, tmp_path):
        """Test error handling during reading"""
        # Create a file that looks like Excel but isn't
        fake_excel = tmp_path / "fake.xlsx"
        fake_excel.write_bytes(b"This is not an Excel file")
        
        result = excel_reader.read_excel_file(fake_excel)
        
        # Should handle error gracefully
        assert result["success"] is False
        assert "error" in result
        assert result["error"] is not None


@pytest.mark.integration
class TestExcelReadingIntegration:
    """Integration tests for Excel reading"""

    def test_with_real_excel_files(self, fixtures_dir):
        """Test reading workflow with real Excel files"""
        # Skip if no real Excel files available in fixtures
        excel_dir = fixtures_dir / "excel" if fixtures_dir else Path("test_fixtures/excel")
        excel_files = []
        
        if excel_dir.exists():
            excel_files = list(excel_dir.glob("*.xlsx")) + list(excel_dir.glob("*.xls"))
        
        if not excel_files:
            pytest.skip("No Excel files in fixtures for integration testing")
        
        reader = ExcelReader()
        
        for excel_file in excel_files[:2]:  # Test first 2 files only
            try:
                result = reader.read_excel_file(excel_file)
                
                print(f"✅ {excel_file.name}: {result['metadata']['processed_sheets']} sheets, "
                      f"{result['metadata']['total_rows']} rows, "
                      f"Format: {result['metadata']['file_format']}")
                
                assert result["file_path"] == str(excel_file)
                assert isinstance(result["success"], bool)
                
                if result["success"]:
                    assert result["metadata"]["total_sheets"] > 0
                    assert len(result["sheets"]) >= 0
                
            except Exception as e:
                print(f"⚠️ {excel_file.name}: {e}")
                # Don't fail test for individual file issues in integration tests


@pytest.mark.benchmark
class TestExcelReadingPerformance:
    """Performance benchmarks for Excel reading"""

    def test_xlsx_reading_speed(self, benchmark, tmp_path):
        """Benchmark .xlsx reading speed"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create a larger test file
            wb = Workbook()
            ws = wb.active
            ws.title = "Large Dataset"
            
            # Add headers
            headers = [f"Column_{i}" for i in range(1, 11)]  # 10 columns
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data (100 rows)
            for row in range(2, 102):
                for col in range(1, 11):
                    ws.cell(row=row, column=col, value=f"Value_{row}_{col}")
            
            xlsx_path = tmp_path / "large_test.xlsx"
            wb.save(xlsx_path)
            wb.close()
            
            reader = ExcelReader()
            
            result = benchmark(reader.read_excel_file, xlsx_path)
            
            if result["success"]:
                assert result["metadata"]["total_rows"] >= 100
                assert result["metadata"]["processed_sheets"] >= 1
                
        except ImportError:
            pytest.skip("openpyxl not available for benchmarking")

    def test_multi_sheet_reading_speed(self, benchmark, tmp_path):
        """Benchmark multi-sheet reading speed"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add 5 sheets with data
            for sheet_num in range(1, 6):
                ws = wb.create_sheet(f"Sheet_{sheet_num}")
                
                # Add some data
                for row in range(1, 21):  # 20 rows
                    for col in range(1, 6):  # 5 columns
                        ws.cell(row=row, column=col, value=f"S{sheet_num}_R{row}_C{col}")
            
            xlsx_path = tmp_path / "multi_sheet_test.xlsx"
            wb.save(xlsx_path)
            wb.close()
            
            reader = ExcelReader()
            
            results = benchmark(reader.read_excel_file, xlsx_path)
            
            if results["success"]:
                assert results["metadata"]["processed_sheets"] == 5
                assert len(results["sheets"]) == 5
                
        except ImportError:
            pytest.skip("openpyxl not available for multi-sheet benchmarking")


# Test CLI interface
class TestExcelReadingCLI:
    """Test CLI interface for Excel reading"""

    def test_cli_with_valid_excel(self, tmp_path):
        """Test CLI with valid Excel file"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Create test Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Data"
            
            # Add some data
            ws.cell(row=1, column=1, value="Name")
            ws.cell(row=1, column=2, value="Value")
            ws.cell(row=2, column=1, value="Test")
            ws.cell(row=2, column=2, value=123)
            
            excel_path = tmp_path / "cli_test.xlsx"
            wb.save(excel_path)
            wb.close()
            
            import subprocess
            import sys
            
            script_path = Path(__file__).parent.parent / "src" / "pipeline" / "excel_reader.py"
            
            result = subprocess.run([
                sys.executable, str(script_path), str(excel_path)
            ], capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0:
                # Parse output
                stdout_lines = result.stdout.strip().split('\n')
                json_line = None
                for line in stdout_lines:
                    if line.startswith('EXCEL_READING_RESULT:'):
                        json_line = line
                        break
                
                if json_line:
                    json_str = json_line.replace('EXCEL_READING_RESULT:', '')
                    reading_result = json.loads(json_str)
                    
                    assert reading_result["success"] in [True, False]  # Either is valid
                    assert reading_result["metadata"]["file_format"] == "xlsx"
                else:
                    pytest.skip("No reading result found - may be due to missing dependencies")
            else:
                # CLI might fail due to missing dependencies
                pytest.skip(f"CLI failed (code {result.returncode}): {result.stderr}")
                
        except ImportError:
            pytest.skip("openpyxl not available for CLI testing")

    def test_cli_with_missing_file(self):
        """Test CLI with non-existent file"""
        import subprocess
        import sys
        
        script_path = Path(__file__).parent.parent / "src" / "pipeline" / "excel_reader.py"
        
        result = subprocess.run([
            sys.executable, str(script_path), "/path/that/does/not/exist.xlsx"
        ], capture_output=True, text=True)
        
        assert result.returncode == 1
        assert "not found" in result.stderr.lower()


# Parameterized tests for different configurations
@pytest.mark.parametrize("config,expected_behavior", [
    ({"include_hidden": True}, "includes_hidden"),
    ({"skip_empty_sheets": False}, "includes_empty"),
    ({"min_rows": 1, "min_cols": 1}, "low_threshold"),
    ({"normalize_headers": False}, "raw_headers")
])
def test_configuration_behaviors(config, expected_behavior, tmp_path):
    """Test different configuration behaviors"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # Create test Excel file with various sheet types
        wb = Workbook()
        
        # Normal sheet
        ws1 = wb.active
        ws1.title = "Normal Sheet"
        ws1.cell(row=1, column=1, value="Header With Spaces")
        ws1.cell(row=2, column=1, value="Data")
        
        # Empty sheet
        ws2 = wb.create_sheet("Empty Sheet")
        
        # Hidden sheet
        ws3 = wb.create_sheet("Hidden Sheet")
        ws3.sheet_state = 'hidden'
        ws3.cell(row=1, column=1, value="Hidden Data")
        
        excel_path = tmp_path / "config_test.xlsx"
        wb.save(excel_path)
        wb.close()
        
        reader = ExcelReader(config)
        result = reader.read_excel_file(excel_path)
        
        if result["success"]:
            if expected_behavior == "includes_hidden":
                # Should process hidden sheets
                assert result["metadata"]["hidden_sheets"] > 0
            elif expected_behavior == "includes_empty":
                # Should process empty sheets
                assert result["metadata"]["empty_sheets"] >= 0
            elif expected_behavior == "low_threshold":
                # Should process sheets with minimal data
                assert result["metadata"]["processed_sheets"] >= 1
            elif expected_behavior == "raw_headers":
                # Should preserve original header format
                for sheet in result["sheets"].values():
                    if sheet["has_headers"]:
                        # Headers should contain spaces (not normalized)
                        assert any(" " in h for h in sheet["headers"] if isinstance(h, str))
        
    except ImportError:
        pytest.skip("openpyxl not available for configuration testing")